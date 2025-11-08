import sys
from pathlib import Path
from openpyxl import load_workbook

path = Path('example_files/Tabelas 2025 - in progress.xlsx')
wb = load_workbook(filename=path, data_only=True)
for ws in wb.worksheets:
    print(f"Sheet: {ws.title}")
    # print first 10 rows limited columns
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=12, values_only=True):
        print('\t', row)
    print('-' * 40)
